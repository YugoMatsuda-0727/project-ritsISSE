import json
import csv
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.chart.series import SeriesLabel

BASE_DIR = Path(__file__).parent
METADATA_DIR = BASE_DIR / "outputs" / "metadata"
DATASET_DIR = BASE_DIR / "outputs" / "dataset"
OUTPUT_PATH = BASE_DIR / "outputs" / "distribution_analysis.xlsx"

SKIP = {".git", "dummy_sample", "Log", "research-scripts"}

def count_classes(metadata_path):
    with open(metadata_path, encoding="utf-8") as f:
        data = json.load(f)
    return len(data)

def get_label_stats(dataset_path):
    total = 0
    positive = 0
    with open(dataset_path, encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            total += 1
            if int(row["co_changed"]) == 1:
                positive += 1
    return total, positive

def collect_data():
    rows = []
    for meta_file in sorted(METADATA_DIR.glob("*.json")):
        project = meta_file.stem
        if project in SKIP:
            continue
        dataset_file = DATASET_DIR / f"{project}_dataset.csv"
        if not dataset_file.exists():
            continue
        class_count = count_classes(meta_file)
        total, positive = get_label_stats(dataset_file)
        ratio = round(positive / total * 100, 2) if total > 0 else 0.0
        rows.append({
            "project": project,
            "class_count": class_count,
            "total_pairs": total,
            "positive_pairs": positive,
            "cochange_ratio": ratio,
        })
        print(f"  {project}: classes={class_count}, ratio={ratio}%")
    return rows

def build_excel(rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Distribution"

    header_fill = PatternFill("solid", start_color="2F4F8F")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    center = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ["Project", "Class数", "ペア合計", "Positive Pairs", "共変更率 (%)"]
    col_widths = [50, 15, 15, 15, 20]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[cell.column_letter].width = w

    ws.row_dimensions[1].height = 22

    alt_fill = PatternFill("solid", start_color="EEF2FF")
    for i, row in enumerate(rows, 2):
        values = [
            row["project"],
            row["class_count"],
            row["total_pairs"],
            row["positive_pairs"],
            row["cochange_ratio"],
        ]
        fill = alt_fill if i % 2 == 0 else None
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.font = Font(name="Arial", size=10)
            cell.alignment = center
            cell.border = border
            if fill:
                cell.fill = fill

    # 相関係数をExcel数式で計算
    last_row = len(rows) + 1
    ws.cell(row=last_row + 2, column=1, value="相関係数（クラス数 vs 共変更率）").font = Font(bold=True, name="Arial")
    corr_cell = ws.cell(row=last_row + 2, column=2)
    corr_cell.value = f"=CORREL(B2:B{last_row},E2:E{last_row})"
    corr_cell.font = Font(name="Arial", color="0000FF")
    corr_cell.number_format = "0.0000"

    # 散布図
    chart = ScatterChart()
    chart.title = "Class Count vs Co-change Rate"
    chart.style = 10
    chart.x_axis.title = "Class Count"
    chart.y_axis.title = "Co-change Rate (%)"
    chart.legend = None

    x_vals = Reference(ws, min_col=2, min_row=2, max_row=last_row)
    y_vals = Reference(ws, min_col=5, min_row=2, max_row=last_row)
    series = Series(y_vals, x_vals, title="Projects")
    series.marker.symbol = "circle"
    series.marker.size = 8
    series.graphicalProperties.line.noFill = True
    chart.series.append(series)
    chart.width = 18
    chart.height = 12

    ws.add_chart(chart, "G2")

    wb.save(OUTPUT_PATH)
    print(f"\n[DONE] {OUTPUT_PATH}")

if __name__ == "__main__":
    print("[COLLECT] データ集計中...")
    rows = collect_data()
    print("\n[BUILD] Excel作成中...")
    build_excel(rows)